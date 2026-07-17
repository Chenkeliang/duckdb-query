from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

import duckdb
from openpyxl import Workbook
import pytest
from core.data.excel_import_manager import cleanup_pending_excel
from fastapi.testclient import TestClient
from main import app

client = TestClient(app)


@pytest.mark.parametrize("header_rows", [1])
def test_excel_upload_inspect_import(tmp_path, header_rows):
    excel_path = Path(tmp_path) / "multi_sheet.xlsx"
    wb = Workbook()
    ws_orders = wb.active
    ws_orders.title = "Orders"
    ws_orders.append(["order_id", "amount"])
    for row in [(101, 12.5), (102, 20.0), (103, 8.75)]:
        ws_orders.append(row)
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["category", "total"])
    for row in [("A", 32.5), ("B", 18.1)]:
        ws_summary.append(row)
    wb.save(excel_path)

    with (
        patch("routers.file_ingestion.schedule_cleanup"),
        patch(
            "core.data.file_datasource_manager.file_datasource_manager.save_file_datasource",
            return_value=True,
        ),
    ):
        with open(excel_path, "rb") as handle:
            upload_resp = client.post(
                "/api/upload",
                data={"table_alias": ""},
                files={
                    "file": (
                        "multi_sheet.xlsx",
                        handle,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

    assert upload_resp.status_code == 200
    upload_body = upload_resp.json()
    upload_data = upload_body.get("data", upload_body)
    assert upload_data.get("pending_excel")
    file_id = upload_data["pending_excel"]["file_id"]

    inspect_resp = client.post(
        "/api/data-sources/excel/inspect", json={"file_id": file_id}
    )
    assert inspect_resp.status_code == 200
    inspect_data = inspect_resp.json()

    inspect_payload = inspect_data.get("data", inspect_data)
    sheets = inspect_payload.get("sheets", [])
    assert len(sheets) == 2
    sheet_names = {sheet["name"] for sheet in sheets}
    assert sheet_names == {"Orders", "Summary"}
    for sheet in sheets:
        assert sheet.get("default_table_name"), (
            "inspect should return default_table_name per sheet"
        )

    payload = {
        "file_id": file_id,
        "sheets": [
            {
                "name": sheet["name"],
                "target_table": sheet.get("default_table_name", sheet["name"]),
                "mode": "replace",
                "header_rows": sheet.get("suggested_header_rows", header_rows),
                "header_row_index": sheet.get("suggested_header_row_index", 1),
                "fill_merged": False,
            }
            for sheet in sheets
        ],
    }

    mem_con = duckdb.connect(":memory:")

    @contextmanager
    def _mem_duckdb():
        yield mem_con

    with (
        patch(
            "core.data.file_datasource_manager.file_datasource_manager.save_file_datasource",
            return_value=True,
        ),
        patch("routers.file_ingestion.with_duckdb_connection", _mem_duckdb),
    ):
        import_resp = client.post("/api/data-sources/excel/import", json=payload)

    assert import_resp.status_code == 200
    import_body = import_resp.json()
    assert import_body.get("success") is True
    import_payload = import_body.get("data", import_body)
    for result in import_payload.get("results", []):
        assert result.get("success") is True
        table_name = result["target_table"]
        assert result.get("row_count", 0) > 0
        count = mem_con.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
        assert count == result["row_count"]
    mem_con.close()
    cleanup_pending_excel(file_id)


def test_excel_sheet_name_with_apostrophe_imports(tmp_path):
    """回归:sheet 名带单引号(如 Q1's Data)曾破坏 read_xlsx('...') SQL,
    仅靠 try/except 回退逐行引擎掩盖。现 escape_string_literal 转义后原生路径正常。"""
    excel_path = Path(tmp_path) / "apos.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1's Data"
    ws.append(["id", "v"])
    for row in [(1, "a"), (2, "b")]:
        ws.append(row)
    wb.save(excel_path)

    with (
        patch("routers.file_ingestion.schedule_cleanup"),
        patch(
            "core.data.file_datasource_manager.file_datasource_manager.save_file_datasource",
            return_value=True,
        ),
    ):
        with open(excel_path, "rb") as handle:
            upload_resp = client.post(
                "/api/upload",
                data={"table_alias": ""},
                files={"file": ("apos.xlsx", handle,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
    file_id = upload_resp.json()["data"]["pending_excel"]["file_id"]
    inspect = client.post("/api/data-sources/excel/inspect", json={"file_id": file_id})
    sheet = inspect.json()["data"]["sheets"][0]
    assert sheet["name"] == "Q1's Data"

    payload = {
        "file_id": file_id,
        "sheets": [{
            "name": "Q1's Data",
            "target_table": "qa_apos",
            "mode": "replace",
            "header_rows": 1,
            "header_row_index": 1,
            "fill_merged": False,
        }],
    }
    mem_con = duckdb.connect(":memory:")

    @contextmanager
    def _mem_duckdb():
        yield mem_con

    with (
        patch("core.data.file_datasource_manager.file_datasource_manager.save_file_datasource",
              return_value=True),
        patch("routers.file_ingestion.with_duckdb_connection", _mem_duckdb),
    ):
        import_resp = client.post("/api/data-sources/excel/import", json=payload)

    body = import_resp.json()
    assert body.get("success") is True, body
    result = body["data"]["results"][0]
    # 撇号 sheet 名不再破坏 SQL,数据完整落库(原生路径转义正确性见
    # test_sql_identifiers.test_neutralizes_read_xlsx_injection)
    assert result["row_count"] == 2
    assert mem_con.execute('SELECT COUNT(*) FROM "qa_apos"').fetchone()[0] == 2
    mem_con.close()
    cleanup_pending_excel(file_id)
