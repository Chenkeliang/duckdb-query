"""calamine 检视分支的合并单元格探测回归。

此前 _inspect_xls_sheets 硬编码 has_merged_cells=False(注释误称 calamine
不暴露合并区信息);实际 python-calamine 0.6.2 提供 merged_cell_ranges。
用 openpyxl 生成带/不带合并区的文件,直接喂给 calamine 检视函数验证。
"""
import os

import pytest
from openpyxl import Workbook

from core.data.excel_import_manager import _inspect_xls_sheets


@pytest.fixture()
def merged_and_plain_workbook(tmp_path):
    wb = Workbook()
    ws_merged = wb.active
    ws_merged.title = "merged"
    ws_merged["A1"] = "header"
    ws_merged["B1"] = "spanning"
    ws_merged["A2"] = 1
    ws_merged["B2"] = 2
    ws_merged.merge_cells("A1:B1")

    ws_plain = wb.create_sheet("plain")
    ws_plain["A1"] = "col"
    ws_plain["A2"] = "value"

    path = os.path.join(tmp_path, "mixed.xlsx")
    wb.save(path)
    return path


def test_calamine_inspect_detects_merged_cells(merged_and_plain_workbook):
    sheets = _inspect_xls_sheets(merged_and_plain_workbook)
    by_name = {s["name"]: s for s in sheets}

    assert by_name["merged"]["has_merged_cells"] is True
    assert by_name["plain"]["has_merged_cells"] is False


def test_calamine_inspect_keeps_row_and_preview_shape(merged_and_plain_workbook):
    # 合并区探测不应影响既有行数/预览结构
    sheets = _inspect_xls_sheets(merged_and_plain_workbook)
    merged = next(s for s in sheets if s["name"] == "merged")
    assert merged["rows"] == 2
    assert merged["columns_count"] == 2
    assert merged["columns"]
