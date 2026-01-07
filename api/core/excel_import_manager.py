import json
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from core.timezone_utils import get_current_time_iso
from core.utils import normalize_dataframe_output

PENDING_BASE_DIR = (
    Path(__file__).resolve().parent.parent / "temp_files" / "excel_pending"
)
PENDING_BASE_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class PendingExcelFile:
    file_id: str
    original_filename: str
    stored_path: str
    uploaded_at: str
    file_size: int
    table_alias: Optional[str]
    default_table_prefix: str


def _metadata_path(file_id: str) -> Path:
    return PENDING_BASE_DIR / file_id / "metadata.json"


def _ensure_unique_name(parts: List[str], index: int) -> str:
    cleaned = [
        p
        for p in [str(part).strip() if part is not None else "" for part in parts]
        if p and p.lower() != "nan"
    ]
    if not cleaned:
        return f"column_{index + 1}"
    candidate = "_".join(cleaned)
    candidate = re.sub(r"[\s]+", "_", candidate, flags=re.UNICODE)
    candidate = re.sub(r"[^\w]", "_", candidate, flags=re.UNICODE).strip("_")
    return candidate or f"column_{index + 1}"


def sanitize_identifier(
    value: str, allow_leading_digit: bool = False, prefix: str = "table"
) -> str:
    if not value:
        value = ""
    sanitized = re.sub(r"[^\w]", "_", value, flags=re.UNICODE)
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    if not sanitized:
        sanitized = f"{prefix}_{uuid4().hex[:8]}"
    if not allow_leading_digit and sanitized[0].isdigit():
        sanitized = f"{prefix}_{sanitized}"
    return sanitized


def register_excel_upload(
    source_path: str, original_filename: str, table_alias: Optional[str] = None
) -> PendingExcelFile:
    file_id = uuid4().hex
    target_dir = PENDING_BASE_DIR / file_id
    target_dir.mkdir(parents=True, exist_ok=True)

    safe_name = original_filename or f"excel_{file_id}.xlsx"
    stored_path = target_dir / safe_name
    shutil.move(source_path, stored_path)

    default_prefix_source = table_alias or Path(safe_name).stem
    default_table_prefix = sanitize_identifier(default_prefix_source or "excel")

    metadata = {
        "file_id": file_id,
        "original_filename": original_filename,
        "stored_path": str(stored_path),
        "uploaded_at": get_current_time_iso(),
        "file_size": os.path.getsize(stored_path),
        "table_alias": table_alias,
        "default_table_prefix": default_table_prefix,
    }

    with _metadata_path(file_id).open("w", encoding="utf-8") as fh:
        json.dump(metadata, fh, ensure_ascii=False, indent=2)

    return PendingExcelFile(**metadata)


def get_pending_excel(file_id: str) -> Optional[PendingExcelFile]:
    metadata_file = _metadata_path(file_id)
    if not metadata_file.exists():
        return None

    with metadata_file.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    stored_path = Path(data.get("stored_path", ""))
    if not stored_path.exists():
        return None

    return PendingExcelFile(**data)


def cleanup_pending_excel(file_id: str):
    target_dir = PENDING_BASE_DIR / file_id
    if target_dir.exists():
        shutil.rmtree(target_dir, ignore_errors=True)


def derive_default_table_name(default_prefix: str, sheet_name: str) -> str:
    sheet_part = sanitize_identifier(
        sheet_name or "sheet", allow_leading_digit=True, prefix="sheet"
    )
    if default_prefix:
        return sanitize_identifier(
            f"{default_prefix}__{sheet_part}", prefix=default_prefix
        )
    return sanitize_identifier(sheet_part, prefix="sheet")


def _map_pandas_dtype_to_duckdb(dtype: Any) -> str:
    dtype_str = str(dtype).lower()
    if "int" in dtype_str:
        return "BIGINT"
    if "float" in dtype_str or "double" in dtype_str:
        return "DOUBLE"
    if "bool" in dtype_str:
        return "BOOLEAN"
    if "datetime" in dtype_str or "date" in dtype_str:
        return "TIMESTAMP"
    if "timedelta" in dtype_str:
        return "INTERVAL"
    if "category" in dtype_str:
        return "VARCHAR"
    return "VARCHAR"


def inspect_excel_sheets(
    file_path: str, preview_rows: int = 20
) -> List[Dict[str, Any]]:
    """检查 Excel 文件的工作表信息，支持 .xlsx 和 .xls 格式"""
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext == ".xls":
        # .xls 文件使用 pandas + xlrd
        return _inspect_xls_sheets(file_path, preview_rows)
    else:
        # .xlsx 文件使用 openpyxl
        return _inspect_xlsx_sheets(file_path, preview_rows)


def _inspect_xlsx_sheets(
    file_path: str, preview_rows: int = 20
) -> List[Dict[str, Any]]:
    """使用 openpyxl 检查 .xlsx 文件"""
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheets_info: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            merged_cells_attr = getattr(sheet, "merged_cells", None)
            if merged_cells_attr is None:
                merged_cells_attr = getattr(sheet, "merged_cell_ranges", None)
            merged = False
            if merged_cells_attr is not None:
                try:
                    merged = bool(
                        getattr(merged_cells_attr, "ranges", merged_cells_attr)
                    )
                except Exception:
                    merged = bool(merged_cells_attr)
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0

            first_row = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])
            ]
            second_row = (
                [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2), [])]
                if max_row >= 2
                else []
            )
            first_empty_ratio = (
                sum(1 for value in first_row if value in (None, "")) / len(first_row)
                if first_row
                else 1.0
            )
            second_empty_ratio = (
                sum(1 for value in second_row if value in (None, "")) / len(second_row)
                if second_row
                else 1.0
            )
            suggested_header_rows = 1
            if merged or (first_empty_ratio > 0.5 and second_empty_ratio < 0.5):
                suggested_header_rows = 2
            suggested_header_row_index = 1
            if first_empty_ratio > 0.5 and second_row:
                suggested_header_row_index = 2

            preview_records: List[Dict[str, Any]] = []
            columns: List[Dict[str, Any]] = []
            try:
                preview_df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    nrows=preview_rows,
                    engine="openpyxl",
                )
                columns = [
                    {
                        "name": str(col),
                        "duckdb_type": _map_pandas_dtype_to_duckdb(dtype),
                    }
                    for col, dtype in preview_df.dtypes.items()
                ]
                preview_records = normalize_dataframe_output(preview_df)
            except Exception:
                columns = []
                preview_records = []

            sheets_info.append(
                {
                    "name": sheet_name,
                    "rows": int(max_row),
                    "columns_count": int(max_col),
                    "has_merged_cells": merged,
                    "suggested_header_rows": suggested_header_rows,
                    "suggested_header_row_index": suggested_header_row_index,
                    "columns": columns,
                    "preview": preview_records,
                }
            )
        return sheets_info
    finally:
        workbook.close()


def _inspect_xls_sheets(file_path: str, preview_rows: int = 20) -> List[Dict[str, Any]]:
    """使用 pandas + xlrd 检查 .xls 文件"""
    import xlrd

    sheets_info: List[Dict[str, Any]] = []
    xl = pd.ExcelFile(file_path, engine="xlrd")

    try:
        for sheet_name in xl.sheet_names:
            # 获取基本信息
            try:
                preview_df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    nrows=preview_rows,
                    engine="xlrd",
                )
                max_col = len(preview_df.columns)
                # 读取全部行数
                full_df = pd.read_excel(file_path, sheet_name=sheet_name, engine="xlrd")
                max_row = len(full_df) + 1  # +1 for header

                columns = [
                    {
                        "name": str(col),
                        "duckdb_type": _map_pandas_dtype_to_duckdb(dtype),
                    }
                    for col, dtype in preview_df.dtypes.items()
                ]
                preview_records = normalize_dataframe_output(preview_df)
            except Exception:
                max_row = 0
                max_col = 0
                columns = []
                preview_records = []

            sheets_info.append(
                {
                    "name": sheet_name,
                    "rows": int(max_row),
                    "columns_count": int(max_col),
                    "has_merged_cells": False,  # xlrd 检测合并单元格较复杂，暂不支持
                    "suggested_header_rows": 1,
                    "suggested_header_row_index": 1,
                    "columns": columns,
                    "preview": preview_records,
                }
            )
        return sheets_info
    finally:
        xl.close()


def _repair_excel_coordinates(file_path: str) -> Optional[str]:
    if not file_path.lower().endswith(".xlsx"):
        return None
    temp_dir = tempfile.mkdtemp(prefix="excel_repair_")
    temp_path = os.path.join(temp_dir, "repaired.xlsx")
    try:
        with zipfile.ZipFile(file_path) as zin, zipfile.ZipFile(temp_path, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith(
                    "xl/worksheets/"
                ) and item.filename.endswith(".xml"):
                    try:
                        text = data.decode("utf-8")
                    except UnicodeDecodeError:
                        pass
                    else:
                        # 增强修复逻辑：
                        # 1. r="1" -> r="A1"
                        # 2. r='"1"' -> r="A1" (quote variations)
                        # 3. Handle potential empty strings if that's the case r=""? No, error says '' is not valid col name.
                        # It likely sees column index '' from coordinate.
                        
                        # 更加宽容的正则，匹配 r="digits"
                        patched = re.sub(r'r=["\'](\d+)["\']', r'r="A\1"', text)
                        data = patched.encode("utf-8")
                zout.writestr(item, data)
    except Exception as e:
        logger.error(f"Excel repair failed: {e}")
        shutil.rmtree(temp_dir, ignore_errors=True)
        return None
    return temp_path


def ensure_unique_columns(names: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for name in names:
        current = name or "column"
        base = current
        if base not in seen:
            seen[base] = 0
            result.append(base)
        else:
            seen[base] += 1
            candidate = f"{base}_{seen[base]}"
            while candidate in seen:
                seen[base] += 1
                candidate = f"{base}_{seen[base]}"
            seen[candidate] = 0
            result.append(candidate)
    return result


def load_excel_sheet_dataframe(
    file_path: str,
    sheet_name: str,
    header_rows: int = 1,
    header_row_index: Optional[int] = 1,
    fill_merged: bool = False,
) -> pd.DataFrame:
    # 根据文件扩展名选择引擎
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext in {".xls"}:
        engine = "xlrd"
    else:
        engine = "openpyxl"

    # 尝试读取 Excel，某些损坏的文件可能需要特殊处理
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            engine=engine,
        )

    except ValueError as e:
        if "is not a valid column name" in str(e):
            logger.warning(f"Openpyxl failed to read {file_path}, trying calamine fallback. Error: {e}")
            # 尝试使用 calamine 引擎（更宽容的解析器）
            fallback_engine = "calamine"
            try:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=None,
                    engine=fallback_engine,
                )
            except Exception as calamine_error:
                logger.error(f"Calamine engine failed: {calamine_error}")
                
                logger.info("Trying experimental XML repair...")
                repair_path = _repair_excel_coordinates(file_path)
                if repair_path:
                    try:
                        df = pd.read_excel(
                            repair_path,
                            sheet_name=sheet_name,
                            header=None,
                            engine="openpyxl",
                        )
                    except Exception as repair_error:
                        logger.error(f"Repaired file read failed: {repair_error}")
                        shutil.rmtree(os.path.dirname(repair_path), ignore_errors=True)
                        raise ValueError(
                            f"Excel 文件已严重损坏。Calamine 引擎报错: {str(calamine_error)}。修复尝试报错: {str(repair_error)}"
                        )
                    else:
                        shutil.rmtree(os.path.dirname(repair_path), ignore_errors=True)
                else:
                    raise ValueError(
                        f"Excel 文件包含无效数据且无法自动修复。建议在 Excel/WPS 中另存为新文件。Calamine 错误: {str(calamine_error)}"
                    )
        else:
            raise

    if fill_merged:
        df = df.ffill(axis=0)

    if header_rows < 0:
        header_rows = 0

    if header_rows == 0:
        headers = [f"column_{idx + 1}" for idx in range(df.shape[1])]
        data_df = df
    else:
        start_index = max((header_row_index or 1) - 1, 0)
        end_index = min(start_index + header_rows, df.shape[0])

        if start_index >= df.shape[0]:
            start_index = 0
            end_index = min(header_rows, df.shape[0])

        header_slice = df.iloc[start_index:end_index]
        drop_range = range(0, end_index)
        data_df = df.drop(index=drop_range, errors="ignore").reset_index(drop=True)
        headers = [
            _ensure_unique_name(
                [
                    header_slice.iloc[row_idx, col_idx]
                    for row_idx in range(header_slice.shape[0])
                    if col_idx < header_slice.shape[1]
                ],
                col_idx,
            )
            for col_idx in range(
                header_slice.shape[1] if header_slice.shape[1] else df.shape[1]
            )
        ]

    headers = [
        sanitize_identifier(name, allow_leading_digit=True, prefix="col")
        for name in headers
    ]
    headers = ensure_unique_columns(headers)
    if len(headers) < data_df.shape[1]:
        headers.extend(
            [f"col_{idx + 1}" for idx in range(len(headers), data_df.shape[1])]
        )
    elif len(headers) > data_df.shape[1]:
        headers = headers[: data_df.shape[1]]
    data_df.columns = headers
    data_df = data_df.dropna(how="all").reset_index(drop=True)

    # 尝试自动推断数值类型
    # 因为初始读取时包含表头，pandas 可能将所有列都设为 object
    # 最稳妥的是尝试 pd.to_numeric
    for col in data_df.columns:
        if data_df[col].dtype == 'object':
            try:
                # 尝试转换为数字，如果整列都是数字（或 NaN）
                numeric_col = pd.to_numeric(data_df[col])
                data_df[col] = numeric_col
            except (ValueError, TypeError):
                # 包含无法转换的文本，保持原样
                pass

    return data_df
