import json
import logging
import os
import re
import shutil
import tempfile
import zipfile

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from datetime import date, datetime, time as dt_time

from openpyxl import load_workbook

from core.common.timezone_utils import get_current_time_iso
from core.common.utils import handle_non_serializable_data

logger = logging.getLogger(__name__)

def _get_pending_base_dir() -> Path:
    from core.common.paths import get_temp_dir

    base = get_temp_dir() / "excel_pending"
    base.mkdir(parents=True, exist_ok=True)
    return base


@dataclass
class PendingExcelFile:
    file_id: str
    original_filename: str
    stored_path: str
    uploaded_at: str
    file_size: int
    table_alias: Optional[str]
    default_table_prefix: str


def _metadata_path(file_id: str) -> Path:
    return _get_pending_base_dir() / file_id / "metadata.json"


def _ensure_unique_name(parts: List[str], index: int) -> str:
    cleaned = [
        p
        for p in [str(part).strip() if part is not None else "" for part in parts]
        if p and p.lower() != "nan"
    ]
    if not cleaned:
        return f"column_{index + 1}"
    candidate = "_".join(cleaned)
    candidate = re.sub(r"[\s]+", "_", candidate, flags=re.UNICODE)
    candidate = re.sub(r"[^\w]", "_", candidate, flags=re.UNICODE).strip("_")
    return candidate or f"column_{index + 1}"


def sanitize_identifier(
    value: str, allow_leading_digit: bool = False, prefix: str = "table"
) -> str:
    if not value:
        value = ""
    sanitized = re.sub(r"[^\w]", "_", value, flags=re.UNICODE)
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    if not sanitized:
        sanitized = f"{prefix}_{uuid4().hex[:8]}"
    if not allow_leading_digit and sanitized[0].isdigit():
        sanitized = f"{prefix}_{sanitized}"
    return sanitized


def register_excel_upload(
    source_path: str, original_filename: str, table_alias: Optional[str] = None
) -> PendingExcelFile:
    file_id = uuid4().hex
    target_dir = _get_pending_base_dir() / file_id
    target_dir.mkdir(parents=True, exist_ok=True)

    safe_name = original_filename or f"excel_{file_id}.xlsx"
    stored_path = target_dir / safe_name
    shutil.move(source_path, stored_path)

    default_prefix_source = table_alias or Path(safe_name).stem
    default_table_prefix = sanitize_identifier(default_prefix_source or "excel")

    metadata = {
        "file_id": file_id,
        "original_filename": original_filename,
        "stored_path": str(stored_path),
        "uploaded_at": get_current_time_iso(),
        "file_size": os.path.getsize(stored_path),
        "table_alias": table_alias,
        "default_table_prefix": default_table_prefix,
    }

    with _metadata_path(file_id).open("w", encoding="utf-8") as fh:
        json.dump(metadata, fh, ensure_ascii=False, indent=2)

    return PendingExcelFile(**metadata)


def get_pending_excel(file_id: str) -> Optional[PendingExcelFile]:
    metadata_file = _metadata_path(file_id)
    if not metadata_file.exists():
        return None

    with metadata_file.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    stored_path = Path(data.get("stored_path", ""))
    if not stored_path.exists():
        return None

    return PendingExcelFile(**data)


def cleanup_pending_excel(file_id: str):
    target_dir = _get_pending_base_dir() / file_id
    if target_dir.exists():
        shutil.rmtree(target_dir, ignore_errors=True)


def derive_default_table_name(default_prefix: str, sheet_name: str) -> str:
    sheet_part = sanitize_identifier(
        sheet_name or "sheet", allow_leading_digit=True, prefix="sheet"
    )
    if default_prefix:
        return sanitize_identifier(
            f"{default_prefix}__{sheet_part}", prefix=default_prefix
        )
    return sanitize_identifier(sheet_part, prefix="sheet")


def _duckdb_type_of_cells(values: List[Any]) -> str:
    """按样本单元格的 Python 类型给出预览用 DuckDB 类型（仅信息展示）。"""
    seen = {type(v) for v in values if v is not None}
    if not seen:
        return "VARCHAR"
    if seen <= {bool}:
        return "BOOLEAN"
    if seen <= {int, bool}:
        return "BIGINT"
    if seen <= {int, float, bool}:
        return "DOUBLE"
    if seen <= {datetime, date}:
        return "TIMESTAMP"
    if seen <= {dt_time}:
        return "TIME"
    return "VARCHAR"


def inspect_excel_sheets(
    file_path: str, preview_rows: int = 20
) -> List[Dict[str, Any]]:
    """检查 Excel 文件的工作表信息，支持 .xlsx 和 .xls 格式"""
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext == ".xls":
        # .xls 文件使用 calamine 原生读取
        return _inspect_xls_sheets(file_path, preview_rows)
    else:
        # .xlsx 文件使用 openpyxl
        return _inspect_xlsx_sheets(file_path, preview_rows)


def _inspect_xlsx_sheets(
    file_path: str, preview_rows: int = 20
) -> List[Dict[str, Any]]:
    """使用 openpyxl 检查 .xlsx 文件"""
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheets_info: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            merged_cells_attr = getattr(sheet, "merged_cells", None)
            if merged_cells_attr is None:
                merged_cells_attr = getattr(sheet, "merged_cell_ranges", None)
            merged = False
            if merged_cells_attr is not None:
                try:
                    merged = bool(
                        getattr(merged_cells_attr, "ranges", merged_cells_attr)
                    )
                except Exception:
                    merged = bool(merged_cells_attr)
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0

            first_row = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])
            ]
            second_row = (
                [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2), [])]
                if max_row >= 2
                else []
            )
            first_empty_ratio = (
                sum(1 for value in first_row if value in (None, "")) / len(first_row)
                if first_row
                else 1.0
            )
            second_empty_ratio = (
                sum(1 for value in second_row if value in (None, "")) / len(second_row)
                if second_row
                else 1.0
            )
            suggested_header_rows = 1
            if merged or (first_empty_ratio > 0.5 and second_empty_ratio < 0.5):
                suggested_header_rows = 2
            suggested_header_row_index = 1
            if first_empty_ratio > 0.5 and second_row:
                suggested_header_row_index = 2

            preview_records: List[Dict[str, Any]] = []
            columns: List[Dict[str, Any]] = []
            try:
                rows_iter = sheet.iter_rows(
                    min_row=1, max_row=preview_rows + 1, values_only=True
                )
                head = [list(r) for r in rows_iter]
                columns, preview_records = _build_preview_from_rows(head)
            except Exception:
                columns = []
                preview_records = []

            sheets_info.append(
                {
                    "name": sheet_name,
                    "rows": int(max_row),
                    "columns_count": int(max_col),
                    "has_merged_cells": merged,
                    "suggested_header_rows": suggested_header_rows,
                    "suggested_header_row_index": suggested_header_row_index,
                    "columns": columns,
                    "preview": preview_records,
                }
            )
        return sheets_info
    finally:
        workbook.close()


def _build_preview_from_rows(head_rows: List[List[Any]]) -> tuple:
    """(表头行+数据行) → (columns 元数据, 预览 records)。首行视作列名。"""
    if not head_rows:
        return [], []
    header = [
        str(v) if v is not None else f"column_{idx + 1}"
        for idx, v in enumerate(head_rows[0])
    ]
    # 重复表头去重(与正式导入 line 515 共用 ensure_unique_columns):columns 与 preview_records
    # 都按 enumerate(header) 位置消费,去重后第 i 个值仍对第 i 个(去重)列名——否则 records 用
    # 原表头作 dict 键,重名列(id,id)后者覆盖前者、首列值丢失(去 pandas 回归,复审 P2)。
    header = ensure_unique_columns(header)
    data_rows = head_rows[1:]
    columns = [
        {
            "name": name,
            "duckdb_type": _duckdb_type_of_cells(
                [row[idx] if idx < len(row) else None for row in data_rows]
            ),
        }
        for idx, name in enumerate(header)
    ]
    preview_records = [
        {
            name: handle_non_serializable_data(
                row[idx] if idx < len(row) else None
            )
            for idx, name in enumerate(header)
        }
        for row in data_rows
    ]
    return columns, preview_records


def _load_sheet_rows_calamine(file_path: str, sheet_name: Optional[str]) -> List[List[Any]]:
    """python-calamine 原生读 sheet 全部行（.xls 主引擎 / .xlsx 兜底引擎）。"""
    from python_calamine import CalamineWorkbook  # pylint: disable=import-error

    workbook = CalamineWorkbook.from_path(file_path)
    target = sheet_name or workbook.sheet_names[0]
    return [list(row) for row in workbook.get_sheet_by_name(target).to_python()]


def _inspect_xls_sheets(file_path: str, preview_rows: int = 20) -> List[Dict[str, Any]]:
    """calamine 原生检查 .xls 文件（v1.2.1 起 xlrd/pandas 退役）。"""
    from python_calamine import CalamineWorkbook  # pylint: disable=import-error

    sheets_info: List[Dict[str, Any]] = []
    workbook = CalamineWorkbook.from_path(file_path)
    for sheet_name in workbook.sheet_names:
        try:
            all_rows = [
                list(row)
                for row in workbook.get_sheet_by_name(sheet_name).to_python()
            ]
            max_row = len(all_rows)
            max_col = max((len(r) for r in all_rows), default=0)
            columns, preview_records = _build_preview_from_rows(
                all_rows[: preview_rows + 1]
            )
        except Exception:
            max_row = 0
            max_col = 0
            columns = []
            preview_records = []

        sheets_info.append(
            {
                "name": sheet_name,
                "rows": int(max_row),
                "columns_count": int(max_col),
                "has_merged_cells": False,  # calamine 不暴露合并区信息，维持原行为
                "suggested_header_rows": 1,
                "suggested_header_row_index": 1,
                "columns": columns,
                "preview": preview_records,
            }
        )
    return sheets_info


def get_excel_native_preview(file_path: str, rows: int = 10) -> Dict[str, Any]:
    """file_utils.get_file_preview 的 Excel 分支：首个 sheet 的原生预览。"""
    sheets = inspect_excel_sheets(file_path, preview_rows=rows)
    first = sheets[0] if sheets else {}
    columns = [col["name"] for col in first.get("columns", [])]
    return {
        "file_type": "excel",
        "file_size": os.path.getsize(file_path),
        "total_rows": max(int(first.get("rows", 0)) - 1, 0),
        "columns": columns,
        "column_types": {
            col["name"]: col["duckdb_type"] for col in first.get("columns", [])
        },
        "preview_data": first.get("preview", [])[:rows],
        "sample_values": {
            name: [
                record[name]
                for record in first.get("preview", [])[:3]
                if record.get(name) is not None
            ]
            for name in columns
        },
    }


def _repair_excel_coordinates(file_path: str) -> Optional[str]:
    if not file_path.lower().endswith(".xlsx"):
        return None
    temp_dir = tempfile.mkdtemp(prefix="excel_repair_")
    temp_path = os.path.join(temp_dir, "repaired.xlsx")
    try:
        with zipfile.ZipFile(file_path) as zin, zipfile.ZipFile(temp_path, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith(
                    "xl/worksheets/"
                ) and item.filename.endswith(".xml"):
                    try:
                        text = data.decode("utf-8")
                    except UnicodeDecodeError:
                        # 增强逻辑：尝试检测编码 (处理国产软件导出的 GBK XML)
                        try:
                            import charset_normalizer
                            matches = charset_normalizer.from_bytes(data).best()
                            if matches:
                                text = data.decode(matches.encoding)
                            else:
                                # 最后的尝试
                                text = data.decode("gb18030", errors="ignore")
                        except Exception:
                            # 实在解不开，跳过此文件
                            continue
                    
                    # 只修 <c> 单元格元素上缺列字母的 r 属性;裸匹配 r="digits"
                    # 会误伤 <row r="N">(行号本就该是纯数字),把整条修复路径打废
                    patched = re.sub(
                        r'(<c\b[^>]*?\br=)["\'](\d+)["\']', r'\1"A\2"', text
                    )
                    
                    # 写回时统一转换为 utf-8
                    data = patched.encode("utf-8")
                zout.writestr(item, data)
    except Exception as e:
        logger.error(f"Excel repair failed: {e}")
        shutil.rmtree(temp_dir, ignore_errors=True)
        return None
    return temp_path


def ensure_unique_columns(names: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for name in names:
        current = name or "column"
        base = current
        if base not in seen:
            seen[base] = 0
            result.append(base)
        else:
            seen[base] += 1
            candidate = f"{base}_{seen[base]}"
            while candidate in seen:
                seen[base] += 1
                candidate = f"{base}_{seen[base]}"
            seen[candidate] = 0
            result.append(candidate)
    return result


def _load_sheet_rows_openpyxl(file_path: str, sheet_name: Optional[str]) -> List[List[Any]]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def load_excel_sheet_rows(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_rows: int = 1,
    header_row_index: Optional[int] = 1,
    fill_merged: bool = False,
) -> tuple:
    """读取 sheet → (列名, 数据行)。纯行式，无 DataFrame。

    读取引擎：.xls 走 calamine；.xlsx 走 openpyxl，坏文件依次退 calamine、
    实验性 XML 修复（与既有行为一致）。单元格保持 Python 原生类型，由下游
    rows_ingest 以忠实文本入库 + 促升引擎定型（与 CSV 同一铁律语义）。
    """
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext == ".xls":
        rows = _load_sheet_rows_calamine(file_path, sheet_name)
    else:
        try:
            rows = _load_sheet_rows_openpyxl(file_path, sheet_name)
        except ValueError as e:
            if "is not a valid column name" not in str(e):
                raise
            logger.warning(
                "Openpyxl failed to read %s, trying calamine fallback. Error: %s",
                file_path, e,
            )
            try:
                rows = _load_sheet_rows_calamine(file_path, sheet_name)
            except Exception as calamine_error:
                logger.error("Calamine engine failed: %s", calamine_error)
                logger.info("Trying experimental XML repair...")
                repair_path = _repair_excel_coordinates(file_path)
                if not repair_path:
                    raise ValueError(
                        "Excel file contains invalid data and cannot be auto-repaired. "
                        f"Please re-save in Excel/WPS. Calamine error: {calamine_error}"
                    ) from calamine_error
                try:
                    rows = _load_sheet_rows_openpyxl(repair_path, sheet_name)
                except Exception as repair_error:
                    raise ValueError(
                        "Excel file is severely corrupted. "
                        f"Calamine engine error: {calamine_error}. "
                        f"Repair attempt error: {repair_error}"
                    ) from repair_error
                finally:
                    shutil.rmtree(os.path.dirname(repair_path), ignore_errors=True)

    if fill_merged:
        # 纵向前向填充（等价原 df.ffill(axis=0)）：合并单元格只有左上格有值
        width = max((len(r) for r in rows), default=0)
        last: List[Any] = [None] * width
        filled: List[List[Any]] = []
        for row in rows:
            padded = list(row) + [None] * (width - len(row))
            for idx in range(width):
                if padded[idx] is None:
                    padded[idx] = last[idx]
                else:
                    last[idx] = padded[idx]
            filled.append(padded)
        rows = filled

    if header_rows < 0:
        header_rows = 0

    width = max((len(r) for r in rows), default=0)

    def _cell(row: List[Any], idx: int) -> Any:
        return row[idx] if idx < len(row) else None

    if header_rows == 0:
        headers = [f"column_{idx + 1}" for idx in range(width)]
        data_rows = rows
    else:
        start_index = max((header_row_index or 1) - 1, 0)
        end_index = min(start_index + header_rows, len(rows))
        if start_index >= len(rows):
            start_index = 0
            end_index = min(header_rows, len(rows))

        header_slice = rows[start_index:end_index]
        data_rows = rows[end_index:]
        headers = [
            _ensure_unique_name(
                [_cell(hrow, col_idx) for hrow in header_slice], col_idx
            )
            for col_idx in range(width)
        ]

    headers = [
        sanitize_identifier(name, allow_leading_digit=True, prefix="col")
        for name in headers
    ]
    headers = ensure_unique_columns(headers)

    # 丢弃全空行（等价原 dropna(how="all")），并按表头宽度补齐/截断
    cleaned_rows = [
        [_cell(row, idx) for idx in range(len(headers))]
        for row in data_rows
        if any(_cell(row, idx) is not None for idx in range(len(headers)))
    ]
    return headers, cleaned_rows
